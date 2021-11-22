
#!/usr/bin/env python
# -*- coding: utf-8 -*-
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A5
from reportlab.lib.units import inch, cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Table
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import PageBreak
from reportlab import *

import openpyxl
import os
from datetime import datetime
from time import sleep

import random
import json
import jsonschema
from jsonschema import validate

step = 0
total_steps = 0
created_files_array = []


def parse_cabins(file_name, config):

    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    cabins_arr = []
    cabin_arr = []
    current = sheet['B3'].value

    # column A = hyttiluokka
    cabin_class_col = config['spreadsheetconfig']['cabin_class_col']

    # column B = hyttinumero
    cabin_id_col = config['spreadsheetconfig']['cabin_id_col']

    # column C = sukunimi
    last_name_col = config['spreadsheetconfig']['last_name_col']

    # column D = etunimi
    first_name_col = config['spreadsheetconfig']['first_name_col']

    # column G = DIN1
    DIN1_col = config['spreadsheetconfig']['DIN1_col']

    # column H = DIN2
    DIN2_col = config['spreadsheetconfig']['DIN2_col']

    # column I = BRE
    BRE_col = config['spreadsheetconfig']['BRE_col']

    # column J = LUN
    LUN_col = config['spreadsheetconfig']['LUN_col']

    # from what row to start the cabins
    first_cabin_row = config['spreadsheetconfig']['first_cabin_row']

    for row in range(first_cabin_row, sheet.max_row + 1):
        # if row is empty, then skip to next row
        if sheet[cabin_class_col+str(row)].value is None and sheet[cabin_id_col+str(row)].value is None and sheet[last_name_col+str(row)].value is None:
            continue
        # if cabin id is the current or there isnt a cabin id -> the same cabin as previous row and add it to the array of this particular cabin
        if sheet[cabin_id_col + str(row)].value == current or sheet[cabin_id_col + str(row)].value is None:
            cabin_arr.append([sheet[cabin_class_col + str(row)].value, sheet[last_name_col + str(row)].value, sheet[first_name_col + str(row)].value,
                              sheet[DIN1_col + str(row)].value, sheet[DIN2_col + str(row)].value, sheet[BRE_col + str(row)].value, sheet[LUN_col + str(row)].value])
        # otherwise the next row is from another cabin -> add the existing cabin to the list of cabins
        else:
            cabins_arr.append(cabin_arr)
            current = sheet[cabin_id_col + str(row)].value
            cabin_arr = [[sheet[cabin_class_col + str(row)].value, sheet[last_name_col + str(row)].value, sheet[first_name_col + str(row)].value,
                          sheet[DIN1_col + str(row)].value, sheet[DIN2_col +
                                                                  str(row)].value, sheet[BRE_col + str(row)].value,
                          sheet[LUN_col + str(row)].value]]
            if row == sheet.max_row:
                cabin_arr.append([sheet[cabin_class_col + str(row)].value, sheet[last_name_col + str(row)].value, sheet[first_name_col + str(row)].value,
                                  sheet[DIN1_col + str(row)].value, sheet[DIN2_col + str(row)].value, sheet[BRE_col + str(row)].value, sheet[LUN_col + str(row)].value])
                cabins_arr.append(cabin_arr)
    return cabins_arr


def create_table(cabin_arr, styleSheet):

    H0 = Paragraph('''<b>Hyttiluokka</b>''', styleSheet["BodyText"])
    H1 = Paragraph('''<b>Sukunimi</b>''', styleSheet["BodyText"])
    H2 = Paragraph('''<b>Etunimi</b>''', styleSheet["BodyText"])
    H3 = Paragraph('''<para align=center><b>I 1</b></para>''',
                   styleSheet["BodyText"])
    H4 = Paragraph('''<para align=center><b>I 2</b></para>''',
                   styleSheet["BodyText"])
    H5 = Paragraph('''<para align=center><b>A</b></para>''',
                   styleSheet["BodyText"])
    H6 = Paragraph('''<para align=center><b>L</b></para>''',
                   styleSheet["BodyText"])

    # TODO: fonts, alignment etc.
    data = []
    header = [H0, H1, H2, H3, H4, H5, H6]

    data.append(header)
    for passenger in cabin_arr:
        data.append(passenger)

    t = Table(data, style=[
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER')
    ])

    t._argW[0] = 3.5 * cm
    for i in range(1, 7):
        if i < 3:
            t._argW[i] = 5 * cm
        else:
            t._argW[i] = 1 * cm
    return t


def createPDF(CABINS_XLSX, OUTPUT_FILE, config):
    step_indicator()

    quote_style = ParagraphStyle('yourtitle',
                                 fontName="Helvetica-Oblique",
                                 fontSize=15,
                                 parent=getSampleStyleSheet()['Heading2'],
                                 alignment=1,
                                 spaceAfter=14,
                                 textColor="red")  # backColor = "rgb(239, 20, 36)",

    # Header image from the config.json

    I = Image(config['picture'])
    I.drawHeight = config['picture_height_on_envelope'] * cm
    I.drawWidth = config['picture_width_on_envelope'] * cm

    cabins = parse_cabins(CABINS_XLSX, config)

    doc = SimpleDocTemplate(OUTPUT_FILE, pagesize=landscape(
        A5), topMargin=1, bottomMargin=0)

    # container for the 'Flowable' objects
    elements = []

    styleSheet = getSampleStyleSheet()
    # Add all elements to doc

    for cabin in cabins:
        # picture
        elements.append(I)

        # artificial padding for bottom with a text element
        elements.append(Paragraph(" ", styleSheet["BodyText"]))

        # table
        elements.append(create_table(cabin, styleSheet))

        # heading
        elements.append(Paragraph(config['heading'], styleSheet["Heading4"]))
        # first note
        elements.append(Paragraph(config['first'], styleSheet["BodyText"]))

        # second note
        elements.append(Paragraph(config['second'], styleSheet["BodyText"]))

        # third note
        elements.append(Paragraph(config['third'], styleSheet["BodyText"]))

        # more artificial padding, TODO read documentation and improve layout to avoid these ugly hacks
        elements.append(Paragraph(" ", styleSheet["BodyText"]))
        elements.append(Paragraph(" ", styleSheet["BodyText"]))

        # random quote
        elements.append(Paragraph(
                        '"'+config['quote'][random.randrange(0, len(config['quote']))]+'"', quote_style))

        elements.append(PageBreak())

    # write the document to disk
    try:
        doc.build(elements)
    except ValueError:
        quit()
    created_files_array.append(doc.filename)
    step_indicator()


def validate_config(config):
    # validate the json schema against the defined schema.json
    try:
        with open('schema.json') as json_file:
            schema = json.load(json_file)
    except ValueError:
        print("schema.json not found and script aborted")
        quit()

    try:
        validate(instance=config, schema=schema)
    except jsonschema.exceptions.ValidationError as err:
        return err
    return 1


def step_indicator(decimals=1):
    global step, total_steps
    percent = ('{0:.' + str(decimals) + 'f}').format(100 *
                                                     ((step+1)/float(total_steps)))
    print(f'\rcreating files {percent}%', end='\r')
    step = step+1
    if step == total_steps:
        print()


def main():

    start_time = datetime.now()

    dirpath = os.getcwd()
    files = []

    try:
        with open('config.json') as json_file:
            config = json.load(json_file)
            if validate_config(config) != 1:
                print("config file invalid:")
                print(validate_config(config))
                quit()
    except ValueError:
        print("config.json not found and script aborted")
        quit()

    # r=root, d=directories, f = files
    for r, d, f in os.walk(dirpath):
        for file in f:
            if file.endswith('.xlsx'):
                files.append(os.path.join(r, file))

    # 2 indication steps per pdf file generation
    global total_steps
    total_steps = len(files)*2

    for xlsx in files:
        path = os.getcwd()
        folder_name = "envelope_print"

        try:
            os.mkdir(folder_name)
        except:
            pass
            # print("envelope_print folder already exists so let's use that one")

        creation_timestamp = datetime.now().strftime("%d.%m.%Y-%H:%M:%S")
        createPDF(
            xlsx, f"{path}/{folder_name}/{folder_name}_{xlsx.split('/')[-1].split('.')[0]}_{creation_timestamp}.pdf", config)

    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    print(f"\ntime elapsed: {duration} seconds")

    print(f"\nfollowing files were read: \n")
    for file in files:
        print(file)
    print(f"\nfollowing files were created: \n")
    for file in created_files_array:
        print(file)


main()
